"""
Region 14 Corrected Subject-Based Mapping Seeder
================================================
This script assigns budget rows to departments based on Subject/Sub-Subject,
not Trustee, and applies Civil micro-segmentation for unmatched items.
"""
import sys
import io
from pathlib import Path
from datetime import datetime
from typing import Dict, Tuple, Optional
from collections import defaultdict

# Fix encoding for Windows console (before any print)
if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8")
    except Exception:
        pass

import argparse
from sqlalchemy import inspect, text
from sqlalchemy.orm import Session

# Ensure project root on path for app imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from app.database import SessionLocal, engine
from app import models
from app.auth_utils import hash_password, verify_password

# Ensure all tables exist
models.Base.metadata.create_all(bind=engine)

# ============================================
# CONFIGURATION
# ============================================

EXCEL_FILE = Path(__file__).parent.parent / "data" / "reports" / "Sarmayei_Region14.xlsx"
REGION_CODE = "14"
REGION_NAME = "منطقه چهارده"
SUBSYSTEM_CODE = "CIVIL_WORKS"
SUBSYSTEM_TITLE = "عمران و طرح‌ها"
DEFAULT_ADMIN_PASSWORD = "Tehran@1403"  # Must be changed on first login

# Departments and Admins
DEPARTMENTS = {
    "GREEN": {
        "code": "R14_GREEN",
        "title": "اداره فضای سبز",
        "org_type": "DEPARTMENT",
        "username": "admin_green_14",
        "full_name": "Admin - Green Space Dept (Region 14)"
    },
    "TRAFFIC": {
        "code": "R14_TRAFFIC",
        "title": "اداره ترافیک",
        "org_type": "DEPARTMENT",
        "username": "admin_traffic_14",
        "full_name": "Admin - Traffic Dept (Region 14)"
    },
    "URBAN": {
        "code": "R14_URBAN",
        "title": "اداره شهرسازی",
        "org_type": "DEPARTMENT",
        "username": "admin_urban_14",
        "full_name": "Admin - Urban Planning Dept (Region 14)"
    },
    "SERVICES": {
        "code": "R14_SERVICES",
        "title": "خدمات شهری/زیباسازی",
        "org_type": "DEPARTMENT",
        "username": "admin_services_14",
        "full_name": "Admin - Beautification & Services (Region 14)"
    }
}

CIVIL_DEPARTMENT = {
    "code": "R14_CIVIL",
    "title": "اداره عمران",
    "org_type": "DEPARTMENT"
}

CIVIL_SECTIONS = {
    "ROAD": {
        "code": "R14_CIVIL_ROAD",
        "title": "راه و آسفالت",
        "org_type": "SECTION"
    },
    "ELEC": {
        "code": "R14_CIVIL_ELEC",
        "title": "تاسیسات برق",
        "org_type": "SECTION"
    },
    "MECH": {
        "code": "R14_CIVIL_MECH",
        "title": "تاسیسات مکانیکی",
        "org_type": "SECTION"
    },
    "SUPERVISION": {
        "code": "R14_CIVIL_SUP",
        "title": "نظارت ابنیه",
        "org_type": "SECTION"
    }
}

# Keyword Logic (Subject-based)
GREEN_KEYWORDS = ["فضای سبز", "پارک"]
TRAFFIC_KEYWORDS = ["ترافیک", "حمل و نقل", "عبور و مرور"]
URBAN_KEYWORDS = ["شهرسازی", "طرح تفصیلی"]
SERVICES_KEYWORDS = ["زیباسازی", "مبلمان شهری", "آبنما"]

# Civil micro-segmentation
CIVIL_ROAD_KEYWORDS = ["آسفالت", "معابر"]
CIVIL_ELEC_KEYWORDS = ["برق", "روشنایی"]
CIVIL_MECH_KEYWORDS = ["آب", "لوله", "تاسیسات"]


# ============================================
# HELPERS
# ============================================

def _is_empty(val) -> bool:
    return val is None or (isinstance(val, float) and val != val)  # NaN check


def clean_str(val) -> Optional[str]:
    if _is_empty(val):
        return None
    s = str(val).strip()
    return s if s else None


def clean_float(val) -> Optional[float]:
    if _is_empty(val):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def normalize_text(val) -> str:
    """Normalize Persian text for keyword matching."""
    s = clean_str(val)
    if not s:
        return ""
    s = s.replace("ي", "ی").replace("ك", "ک")
    return s.lower()


def contains_any(text: str, keywords) -> bool:
    return any(keyword in text for keyword in keywords)


def get_first_value(row: dict, columns) -> Optional[str]:
    for col in columns:
        if col in row and not _is_empty(row.get(col)):
            return row[col]
    return None


def load_excel_rows(excel_path: Path) -> list:
    """Load first sheet of Excel file as list of dicts (column header -> value). Uses openpyxl only."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    header_row = list(ws.iter_rows(min_row=1, max_row=1))
    if not header_row:
        wb.close()
        return []
    headers = []
    for i, c in enumerate(header_row[0], 1):
        v = c.value
        headers.append(str(v).strip() if v is not None and str(v).strip() else f"_col{i}")
    rows = []
    for row in ws.iter_rows(min_row=2):
        rows.append(dict(zip(headers, (c.value for c in row))))
    wb.close()
    return rows


# ============================================
# DATABASE OPERATIONS
# ============================================

_ACTIVITY_HAS_ORG_UNIT = None


def activity_org_unit_supported(db: Session) -> bool:
    """Check if subsystem_activities has org_unit_id column."""
    global _ACTIVITY_HAS_ORG_UNIT
    if _ACTIVITY_HAS_ORG_UNIT is None:
        inspector = inspect(db.get_bind())
        columns = {col["name"] for col in inspector.get_columns("subsystem_activities")}
        _ACTIVITY_HAS_ORG_UNIT = "org_unit_id" in columns
    return _ACTIVITY_HAS_ORG_UNIT


def get_or_create_region(db: Session) -> int:
    region = db.query(models.OrgUnit).filter(
        models.OrgUnit.code == REGION_CODE,
        models.OrgUnit.org_type == "ZONE"
    ).first()
    if region:
        return region.id

    region = models.OrgUnit(
        title=REGION_NAME,
        code=REGION_CODE,
        org_type="ZONE"
    )
    db.add(region)
    db.flush()
    print(f"  [CREATED] Region: {REGION_NAME} (ID: {region.id})")
    return region.id


def get_or_create_subsystem(db: Session) -> int:
    subsystem = db.query(models.Subsystem).filter(
        models.Subsystem.code == SUBSYSTEM_CODE
    ).first()
    if subsystem:
        return subsystem.id

    subsystem = models.Subsystem(
        code=SUBSYSTEM_CODE,
        title=SUBSYSTEM_TITLE,
        icon="construction",
        is_active=True,
        order=14,
        attachment_type="upload"
    )
    db.add(subsystem)
    db.flush()
    print(f"  [CREATED] Subsystem: {SUBSYSTEM_TITLE} (ID: {subsystem.id})")
    return subsystem.id


def get_or_create_org_unit(
    db: Session,
    title: str,
    code: str,
    parent_id: Optional[int],
    org_type: str
) -> Tuple[int, bool]:
    existing = db.query(models.OrgUnit).filter(
        models.OrgUnit.code == code
    ).first()
    if existing:
        return existing.id, False

    unit = models.OrgUnit(
        title=title,
        code=code,
        parent_id=parent_id,
        org_type=org_type
    )
    db.add(unit)
    db.flush()
    return unit.id, True


def ensure_section_subsystem_access(db: Session, section_id: int, subsystem_id: int) -> bool:
    existing = db.query(models.SectionSubsystemAccess).filter(
        models.SectionSubsystemAccess.section_id == section_id,
        models.SectionSubsystemAccess.subsystem_id == subsystem_id
    ).first()
    if existing:
        return False
    access = models.SectionSubsystemAccess(
        section_id=section_id,
        subsystem_id=subsystem_id
    )
    db.add(access)
    db.flush()
    return True


def create_or_update_admin_user(
    db: Session,
    username: str,
    full_name: str,
    dept_id: int,
    zone_id: int,
    subsystem_id: int
) -> Tuple[int, bool, bool]:
    existing = db.query(models.User).filter(
        models.User.username == username
    ).first()

    if existing:
        updated = False

        if not verify_password(DEFAULT_ADMIN_PASSWORD, existing.password_hash or ""):
            existing.password_hash = hash_password(DEFAULT_ADMIN_PASSWORD)
            updated = True

        if existing.role != "ADMIN_L2":
            existing.role = "ADMIN_L2"
            updated = True
        if existing.admin_level != 2:
            existing.admin_level = 2
            updated = True
        if existing.default_dept_id != dept_id:
            existing.default_dept_id = dept_id
            updated = True
        if existing.default_zone_id != zone_id:
            existing.default_zone_id = zone_id
            updated = True
        if existing.default_section_id is not None:
            existing.default_section_id = None
            updated = True
        if existing.managed_subsystem_ids is not None:
            existing.managed_subsystem_ids = None
            updated = True
        if not existing.is_active:
            existing.is_active = True
            updated = True

        # Ensure subsystem access exists
        access_exists = db.query(models.UserSubsystemAccess).filter(
            models.UserSubsystemAccess.user_id == existing.id,
            models.UserSubsystemAccess.subsystem_id == subsystem_id
        ).first()
        if not access_exists:
            access = models.UserSubsystemAccess(
                user_id=existing.id,
                subsystem_id=subsystem_id
            )
            db.add(access)
            updated = True

        if updated:
            db.flush()

        return existing.id, False, updated

    user = models.User(
        username=username,
        password_hash=hash_password(DEFAULT_ADMIN_PASSWORD),
        full_name=full_name,
        role="ADMIN_L2",
        admin_level=2,
        default_zone_id=zone_id,
        default_dept_id=dept_id,
        managed_subsystem_ids=None,
        is_active=True
    )
    db.add(user)
    db.flush()

    access = models.UserSubsystemAccess(
        user_id=user.id,
        subsystem_id=subsystem_id
    )
    db.add(access)
    db.flush()

    return user.id, True, False


def create_unique_activity(
    db: Session,
    subsystem_id: int,
    budget_code: str,
    description: str,
    row_index: int,
    activity_scope: str,
    org_unit_id: Optional[int] = None
) -> Tuple[int, bool]:
    activity_code = f"R14_{activity_scope}_{budget_code}_{row_index}"
    existing = db.query(models.SubsystemActivity).filter(
        models.SubsystemActivity.code == activity_code
    ).first()
    if existing:
        return existing.id, False

    activity = models.SubsystemActivity(
        subsystem_id=subsystem_id,
        code=activity_code,
        title=description[:200] if description else f"Budget {budget_code}",
        is_active=True,
        frequency=models.ActivityFrequency.YEARLY,
        requires_file_upload=True
    )
    db.add(activity)
    db.flush()

    if org_unit_id and activity_org_unit_supported(db):
        db.execute(
            text(
                "UPDATE subsystem_activities SET org_unit_id = :org_unit_id WHERE id = :activity_id"
            ),
            {"org_unit_id": org_unit_id, "activity_id": activity.id}
        )

    return activity.id, True


def create_budget_row(
    db: Session,
    activity_id: int,
    budget_code: str,
    description: str,
    approved_amount: float,
    org_unit_id: int
) -> Tuple[int, bool]:
    existing = db.query(models.BudgetRow).filter(
        models.BudgetRow.budget_coding == budget_code,
        models.BudgetRow.org_unit_id == org_unit_id
    ).first()
    if existing:
        return existing.id, False

    approved = int(approved_amount) if approved_amount else 0

    budget_row = models.BudgetRow(
        activity_id=activity_id,
        org_unit_id=org_unit_id,
        budget_coding=budget_code,
        description=description[:500] if description else "",
        approved_amount=approved,
        blocked_amount=0,
        spent_amount=0,
        fiscal_year="1403"
    )
    db.add(budget_row)
    db.flush()
    return budget_row.id, True


def create_activity_constraint(
    db: Session,
    activity_id: int,
    budget_code: str
) -> Tuple[int, bool]:
    existing = db.query(models.ActivityConstraint).filter(
        models.ActivityConstraint.subsystem_activity_id == activity_id,
        models.ActivityConstraint.budget_code_pattern == budget_code
    ).first()
    if existing:
        return existing.id, False

    constraint = models.ActivityConstraint(
        subsystem_activity_id=activity_id,
        budget_code_pattern=budget_code,
        constraint_type="INCLUDE",
        description=f"Subject-based mapping: Activity locked to budget {budget_code}",
        is_active=True,
        priority=100
    )
    db.add(constraint)
    db.flush()
    return constraint.id, True


# ============================================
# CLASSIFICATION LOGIC
# ============================================

def classify_department(subject: str, sub_subject: str, description: str) -> Tuple[str, Optional[str]]:
    subject_text = normalize_text(subject)
    sub_subject_text = normalize_text(sub_subject)
    desc_text = normalize_text(description)

    if contains_any(subject_text, GREEN_KEYWORDS):
        return "GREEN", None
    if contains_any(subject_text, TRAFFIC_KEYWORDS):
        return "TRAFFIC", None
    if contains_any(subject_text, URBAN_KEYWORDS):
        return "URBAN", None
    if contains_any(subject_text, SERVICES_KEYWORDS):
        return "SERVICES", None

    # Civil micro-segmentation
    if contains_any(sub_subject_text, CIVIL_ROAD_KEYWORDS):
        return "CIVIL", "ROAD"
    if contains_any(desc_text, CIVIL_ELEC_KEYWORDS):
        return "CIVIL", "ELEC"
    if contains_any(desc_text, CIVIL_MECH_KEYWORDS):
        return "CIVIL", "MECH"
    return "CIVIL", "SUPERVISION"


# ============================================
# MAIN SEEDING LOGIC
# ============================================

def seed_region14_corrected(db: Session, rows: list, dry_run: bool = False) -> Dict:
    stats = {
        "total_rows": 0,
        "org_units_created": 0,
        "users_created": 0,
        "users_updated": 0,
        "section_access_created": 0,
        "activities_created": 0,
        "budget_rows_created": 0,
        "constraints_created": 0,
        "skipped": 0,
        "errors": [],
        "assignments": defaultdict(int)
    }

    print("\n" + "=" * 80)
    print("REGION 14 CORRECTED SUBJECT-BASED SEEDER")
    print("=" * 80)
    print(f"Mode: {'DRY RUN (Preview)' if dry_run else 'LIVE (Committing)'}")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)

    stats["total_rows"] = len(rows)

    # Step 1: Ensure parent structures
    print("\n[1/6] Creating parent structures...")
    region_id = get_or_create_region(db)
    subsystem_id = get_or_create_subsystem(db)

    # Step 2: Create departments and civil sections
    print("\n[2/6] Creating OrgUnits (Departments + Civil Sections)...")
    dept_ids = {}

    # Civil Department (parent for sections)
    civil_dept_id, created = get_or_create_org_unit(
        db,
        CIVIL_DEPARTMENT["title"],
        CIVIL_DEPARTMENT["code"],
        region_id,
        CIVIL_DEPARTMENT["org_type"]
    )
    dept_ids["CIVIL"] = civil_dept_id
    if created:
        stats["org_units_created"] += 1
        print(f"  [NEW] {CIVIL_DEPARTMENT['title']} (ID: {civil_dept_id})")
    else:
        print(f"  [EXISTS] {CIVIL_DEPARTMENT['title']} (ID: {civil_dept_id})")

    # Civil Sections
    civil_section_ids = {}
    for section_key, config in CIVIL_SECTIONS.items():
        section_id, created = get_or_create_org_unit(
            db,
            config["title"],
            config["code"],
            civil_dept_id,
            config["org_type"]
        )
        civil_section_ids[section_key] = section_id
        if created:
            stats["org_units_created"] += 1
            print(f"  [NEW] Civil {config['title']} (ID: {section_id})")
        else:
            print(f"  [EXISTS] Civil {config['title']} (ID: {section_id})")

    # Other Departments
    for dept_key, config in DEPARTMENTS.items():
        dept_id, created = get_or_create_org_unit(
            db,
            config["title"],
            config["code"],
            region_id,
            config["org_type"]
        )
        dept_ids[dept_key] = dept_id
        if created:
            stats["org_units_created"] += 1
            print(f"  [NEW] {config['title']} (ID: {dept_id})")
        else:
            print(f"  [EXISTS] {config['title']} (ID: {dept_id})")

    # Step 3: Create admin users
    print("\n[3/6] Creating admin users for departments...")
    for dept_key, config in DEPARTMENTS.items():
        dept_id = dept_ids[dept_key]
        user_id, created, updated = create_or_update_admin_user(
            db,
            config["username"],
            config["full_name"],
            dept_id,
            region_id,
            subsystem_id
        )
        if created:
            stats["users_created"] += 1
            print(f"  [NEW] {config['username']} (ID: {user_id})")
            print(f"        Password: {DEFAULT_ADMIN_PASSWORD}")
        elif updated:
            stats["users_updated"] += 1
            print(f"  [UPDATED] {config['username']} (ID: {user_id})")
            print(f"        Password: {DEFAULT_ADMIN_PASSWORD}")
        else:
            print(f"  [EXISTS] {config['username']} (ID: {user_id})")

    # Ensure section-to-subsystem mapping exists (for all OrgUnits)
    all_org_units = list(dept_ids.values()) + list(civil_section_ids.values())
    for org_id in all_org_units:
        if ensure_section_subsystem_access(db, org_id, subsystem_id):
            stats["section_access_created"] += 1

    # Step 4: Process each budget row
    print("\n[4/6] Processing budget rows (Subject-based mapping)...")
    for idx, row in enumerate(rows):
        budget_code = None
        try:
            budget_code = clean_str(get_first_value(row, ["کد بودجه", "کد"]))
            if not budget_code:
                stats["skipped"] += 1
                continue

            description = clean_str(get_first_value(row, ["شرح ردیف", "شرح"]))
            approved_amount = clean_float(get_first_value(row, ["مصوب 1403", "مصوب"]))
            subject = clean_str(get_first_value(row, ["موضوع"]))
            sub_subject = clean_str(get_first_value(row, ["زیر موضوع", "زیرموضوع"]))

            dept_key, section_key = classify_department(subject, sub_subject, description)

            if dept_key == "CIVIL":
                org_unit_id = civil_section_ids[section_key]
                assignment_key = f"CIVIL_{section_key}"
                activity_scope = assignment_key
            else:
                org_unit_id = dept_ids[dept_key]
                assignment_key = dept_key
                activity_scope = dept_key

            stats["assignments"][assignment_key] += 1

            activity_id, activity_created = create_unique_activity(
                db,
                subsystem_id,
                budget_code,
                description,
                idx,
                activity_scope,
                org_unit_id=org_unit_id
            )

            budget_row_id, budget_created = create_budget_row(
                db,
                activity_id,
                budget_code,
                description,
                approved_amount,
                org_unit_id
            )

            constraint_id, constraint_created = create_activity_constraint(
                db,
                activity_id,
                budget_code
            )

            if activity_created:
                stats["activities_created"] += 1
            if budget_created:
                stats["budget_rows_created"] += 1
            if constraint_created:
                stats["constraints_created"] += 1

            if (idx + 1) % 50 == 0:
                print(f"  Processed {idx + 1}/{len(rows)} rows...")

        except Exception as e:
            error_msg = f"Row {idx} (Budget: {budget_code}): {str(e)}"
            stats["errors"].append(error_msg)
            print(f"  [ERROR] {error_msg}")

    # Step 5: Finalize
    print("\n[5/6] Finalizing...")
    if dry_run:
        print("  [DRY RUN] Rolling back all changes...")
        db.rollback()
    else:
        print("  [COMMIT] Saving all changes to database...")
        db.commit()

    return stats


def print_summary(stats: Dict, dry_run: bool):
    print("\n" + "=" * 80)
    print("SEEDING SUMMARY REPORT")
    print("=" * 80)
    print(f"Status: {'DRY RUN COMPLETED' if dry_run else 'IMPORT COMPLETED'}")
    print()
    print("Structure:")
    print(f"  • OrgUnits Created: {stats['org_units_created']}")
    print(f"  • Admin Users Created: {stats['users_created']}")
    print(f"  • Admin Users Updated: {stats['users_updated']}")
    print(f"  • Section Access Links Created: {stats['section_access_created']}")
    print()
    print("Budget Components:")
    print(f"  • Total Rows Processed: {stats['total_rows']}")
    print(f"  • Activities Created: {stats['activities_created']}")
    print(f"  • BudgetRows Created: {stats['budget_rows_created']}")
    print(f"  • ActivityConstraints Created: {stats['constraints_created']}")
    print(f"  • Skipped: {stats['skipped']}")
    print()

    assignment = stats["assignments"]
    print("Assignment Summary:")
    print(
        "  Assigned {green} items to Green Space, {traffic} to Traffic, {urban} to Urban Planning, "
        "{services} to Beautification/Services, {c_road} to Civil-Road, {c_elec} to Civil-Elec, "
        "{c_mech} to Civil-Mech, {c_sup} to Civil-Supervision."
        .format(
            green=assignment.get("GREEN", 0),
            traffic=assignment.get("TRAFFIC", 0),
            urban=assignment.get("URBAN", 0),
            services=assignment.get("SERVICES", 0),
            c_road=assignment.get("CIVIL_ROAD", 0),
            c_elec=assignment.get("CIVIL_ELEC", 0),
            c_mech=assignment.get("CIVIL_MECH", 0),
            c_sup=assignment.get("CIVIL_SUPERVISION", 0)
        )
    )
    print()

    if stats["errors"]:
        print(f"⚠ Errors Encountered: {len(stats['errors'])}")
        for error in stats["errors"][:5]:
            print(f"  - {error}")
        if len(stats["errors"]) > 5:
            print(f"  ... and {len(stats['errors']) - 5} more")
        print()

    print("=" * 80)
    print("Department Admin Credentials:")
    print("=" * 80)
    for dept_key, config in DEPARTMENTS.items():
        print(f"  • {config['username']}")
        print(f"    Name: {config['full_name']}")
        print(f"    Password: {DEFAULT_ADMIN_PASSWORD} (MUST CHANGE)")
    print("=" * 80)


# ============================================
# CLI ENTRY POINT
# ============================================

def main():
    parser = argparse.ArgumentParser(
        description="Region 14 Corrected Subject-Based Seeder",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Preview what will be created
  python seed_region14_corrected.py --dry-run

  # Execute the import with default Excel location
  python seed_region14_corrected.py

  # Specify a custom Excel file
  python seed_region14_corrected.py --excel path/to/Sarmayei_Region14.xlsx
        """
    )
    parser.add_argument(
        "--excel",
        type=str,
        default=str(EXCEL_FILE),
        help=f"Path to Excel file (default: {EXCEL_FILE})"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview changes without committing to database"
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"\n[ERROR] Excel file not found: {excel_path}")
        return 1

    print("=" * 80)
    print("REGION 14 SUBJECT-BASED MAPPING IMPORT")
    print("=" * 80)
    print(f"Execution Mode: {'DRY RUN' if args.dry_run else 'LIVE IMPORT'}")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 80)

    try:
        rows = load_excel_rows(excel_path)
        n_cols = len(rows[0]) if rows else 0
        print(f"\n[LOAD] Loaded {len(rows)} rows, {n_cols} columns from {excel_path.name}")
    except Exception as e:
        print(f"\n[ERROR] Failed to read Excel file: {e}")
        return 1

    db = SessionLocal()
    try:
        stats = seed_region14_corrected(db, rows, dry_run=args.dry_run)
        print_summary(stats, args.dry_run)
        return 0
    except Exception as e:
        print("\n" + "=" * 80)
        print("[FATAL ERROR] Seeding Failed")
        print("=" * 80)
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        db.rollback()
        return 1
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
