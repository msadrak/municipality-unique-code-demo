"""
Export Final Clean Report - V6 Config to Excel
================================================
Converts app/config/config_master.json (V6) to a polished Excel report.

Fixes:
- Correctly parses budget type from constraints[0].allowed_budget_types
- RTL sheet direction
- Auto-adjusted column widths

Usage:
    python scripts/export_final_clean_report.py
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = "app/config/config_master.json"
OUTPUT_FILE = "لیست_نهایی_فعالیت_ها_V6.xlsx"

# Translation mappings
BUDGET_TYPE_MAP = {
    "capital": "عمرانی (سرمایه‌ای)",
    "expense": "جاری (هزینه‌ای)",
}

FREQUENCY_MAP = {
    "MONTHLY": "مستمر",
    "DAILY": "مستمر",
    "WEEKLY": "مستمر",
    "YEARLY": "موردی",
}


# ============================================================
# DATA PROCESSING
# ============================================================

def load_config() -> dict:
    """Load the JSON config file."""
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def parse_budget_type(constraints: list) -> str:
    """
    CRITICAL FIX: Correctly parse budget type from constraints.
    Path: constraints[0]['allowed_budget_types']
    """
    if not constraints:
        return "عمومی"
    
    try:
        # Get allowed_budget_types from first constraint
        first_constraint = constraints[0]
        allowed_types = first_constraint.get("allowed_budget_types", [])
        
        if not allowed_types:
            return "عمومی"
        
        has_capital = "capital" in allowed_types
        has_expense = "expense" in allowed_types
        
        if has_capital and has_expense:
            return "عمومی"
        elif has_capital:
            return BUDGET_TYPE_MAP["capital"]
        elif has_expense:
            return BUDGET_TYPE_MAP["expense"]
        else:
            return "عمومی"
            
    except (KeyError, IndexError, TypeError):
        return "عمومی"


def parse_frequency(frequency: str) -> str:
    """Translate frequency to Persian."""
    if not frequency:
        return "موردی"
    return FREQUENCY_MAP.get(frequency.upper(), "موردی")


def flatten_config(config: dict) -> list:
    """Flatten the nested JSON into a list of rows for Excel."""
    rows = []
    
    for subsystem in config.get("subsystems", []):
        system_name = subsystem.get("title", "")
        
        for activity in subsystem.get("activities", []):
            activity_code = activity.get("code", "")
            activity_title = activity.get("title", "")
            frequency = activity.get("frequency", "")
            constraints = activity.get("constraints", [])
            
            # CRITICAL: Parse budget type correctly
            budget_type = parse_budget_type(constraints)
            nature = parse_frequency(frequency)
            
            rows.append({
                "نام سامانه": system_name,
                "عنوان فعالیت": activity_title,
                "نوع بودجه": budget_type,
                "ماهیت": nature,
                "کد سیستمی": activity_code,
                "تایید حسابدار": "",  # Empty for manual input
                "توضیحات": ""  # Empty for manual input
            })
    
    return rows


# ============================================================
# EXCEL FORMATTING
# ============================================================

def style_excel(output_path: str):
    """Apply professional RTL styling to the Excel file."""
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alternating row colors
    light_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Style data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = light_fill
            else:
                cell.fill = white_fill
    
    # Auto-adjust column widths
    column_widths = {
        'A': 32,  # نام سامانه
        'B': 45,  # عنوان فعالیت
        'C': 22,  # نوع بودجه
        'D': 12,  # ماهیت
        'E': 20,  # کد سیستمی
        'F': 15,  # تایید حسابدار
        'G': 25   # توضیحات
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row heights
    ws.row_dimensions[1].height = 30  # Header
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Set sheet direction to RTL (Right-to-Left for Persian)
    ws.sheet_view.rightToLeft = True
    
    wb.save(output_path)


# ============================================================
# MAIN
# ============================================================

def main():
    """Main entry point."""
    print("=" * 65)
    print("📊 EXPORT FINAL CLEAN REPORT - V6")
    print("=" * 65)
    
    # Load config
    print(f"\n📁 Loading: {INPUT_FILE}")
    try:
        config = load_config()
    except FileNotFoundError:
        print(f"❌ Error: File not found: {INPUT_FILE}")
        return
    except json.JSONDecodeError as e:
        print(f"❌ Error: Invalid JSON: {e}")
        return
    
    # Flatten data
    print("🔄 Processing data...")
    rows = flatten_config(config)
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Verify budget type parsing
    budget_counts = df["نوع بودجه"].value_counts()
    print(f"\n📋 Budget Type Distribution:")
    for bt, count in budget_counts.items():
        print(f"   • {bt}: {count}")
    
    # Summary
    print(f"\n📊 Summary:")
    print(f"   Config Version: {config.get('version', 'N/A')}")
    print(f"   Total Subsystems: {len(config.get('subsystems', []))}")
    print(f"   Total Activities: {len(rows)}")
    
    # Export to Excel
    print(f"\n💾 Exporting to: {OUTPUT_FILE}")
    df.to_excel(OUTPUT_FILE, index=False, engine='openpyxl')
    
    # Apply styling
    print("🎨 Applying RTL styling...")
    style_excel(OUTPUT_FILE)
    
    print(f"\n✅ Done! File saved: {OUTPUT_FILE}")
    print("   📝 Columns for Accountant Review:")
    print("      • 'تایید حسابدار' - Approval")
    print("      • 'توضیحات' - Notes")
    print("=" * 65)


if __name__ == "__main__":
    main()
