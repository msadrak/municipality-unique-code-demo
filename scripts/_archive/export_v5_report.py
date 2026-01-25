"""
Export V5 Config to Persian Excel Report
==========================================
Converts app/config/config_master_v5.json to a professional Excel report
for Accounting Department verification.

Usage:
    python scripts/export_v5_report.py
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = "app/config/config_master_v5.json"
OUTPUT_FILE = "گزارش_تایید_نهایی_V5.xlsx"

# Translation mappings
BUDGET_TYPE_TRANSLATIONS = {
    "capital": "عمرانی (سرمایه‌ای)",
    "expense": "جاری (هزینه‌ای)",
}

FREQUENCY_TRANSLATIONS = {
    "MONTHLY": "مستمر",
    "DAILY": "مستمر",
    "WEEKLY": "مستمر",
    "YEARLY": "موردی",
}


# ============================================================
# DATA PROCESSING
# ============================================================

def load_config() -> dict:
    """Load the JSON config file."""
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_budget_type(constraints: list) -> str:
    """Extract and translate budget type from constraints."""
    if not constraints:
        return "نامشخص"
    
    try:
        allowed = constraints[0].get("allowed_budget_types", [])
        if not allowed:
            return "نامشخص"
        
        if "capital" in allowed and "expense" in allowed:
            return "هر دو"
        elif "capital" in allowed:
            return BUDGET_TYPE_TRANSLATIONS["capital"]
        elif "expense" in allowed:
            return BUDGET_TYPE_TRANSLATIONS["expense"]
        else:
            return "نامشخص"
    except (KeyError, IndexError, TypeError):
        return "نامشخص"


def get_frequency(frequency: str) -> str:
    """Translate frequency to Persian."""
    if not frequency:
        return "موردی"
    return FREQUENCY_TRANSLATIONS.get(frequency.upper(), "موردی")


def flatten_config(config: dict) -> list:
    """Flatten the nested JSON into a list of rows for Excel."""
    rows = []
    
    for subsystem in config.get("subsystems", []):
        system_name = subsystem.get("title", "")
        
        for activity in subsystem.get("activities", []):
            activity_code = activity.get("code", "")
            activity_title = activity.get("title", "")
            frequency = activity.get("frequency", "")
            constraints = activity.get("constraints", [])
            
            budget_type = get_budget_type(constraints)
            nature = get_frequency(frequency)
            
            rows.append({
                "نام سامانه": system_name,
                "عنوان فعالیت": activity_title,
                "نوع بودجه": budget_type,
                "ماهیت": nature,
                "کد سیستمی": activity_code,
                "تایید حسابدار": "",  # Empty for manual input
                "توضیحات اصلاحی": ""  # Empty for manual input
            })
    
    return rows


# ============================================================
# EXCEL FORMATTING
# ============================================================

def style_excel(output_path: str):
    """Apply professional styling to the Excel file."""
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alternating row colors
    light_fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Style data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = thin_border
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = light_fill
            else:
                cell.fill = white_fill
    
    # Auto-adjust column widths
    column_widths = {
        'A': 30,  # نام سامانه
        'B': 40,  # عنوان فعالیت
        'C': 22,  # نوع بودجه
        'D': 12,  # ماهیت
        'E': 18,  # کد سیستمی
        'F': 15,  # تایید حسابدار
        'G': 25   # توضیحات اصلاحی
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height
    ws.row_dimensions[1].height = 30  # Header
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 22
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Set sheet direction to RTL (Right-to-Left for Persian)
    ws.sheet_view.rightToLeft = True
    
    wb.save(output_path)


# ============================================================
# MAIN
# ============================================================

def main():
    """Main entry point."""
    print("=" * 60)
    print("📊 گزارش‌ساز V5 - Export to Excel")
    print("=" * 60)
    
    # Load config
    print(f"\n📁 Loading: {INPUT_FILE}")
    try:
        config = load_config()
    except FileNotFoundError:
        print(f"❌ Error: File not found: {INPUT_FILE}")
        print("   Please run generate_config_v5_full.py first.")
        return
    except json.JSONDecodeError as e:
        print(f"❌ Error: Invalid JSON: {e}")
        return
    
    # Flatten data
    print("🔄 Processing data...")
    rows = flatten_config(config)
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Calculate summary
    print(f"\n📋 Summary:")
    print(f"   Config Version: {config.get('version', 'N/A')}")
    print(f"   Total Subsystems: {len(config.get('subsystems', []))}")
    print(f"   Total Activities: {len(rows)}")
    
    # Group by subsystem
    print("\n   Activities per Subsystem:")
    for subsystem in config.get("subsystems", []):
        act_count = len(subsystem.get("activities", []))
        print(f"   • {subsystem.get('title', '')}: {act_count}")
    
    # Export to Excel
    print(f"\n💾 Exporting to: {OUTPUT_FILE}")
    df.to_excel(OUTPUT_FILE, index=False, engine='openpyxl')
    
    # Apply styling
    print("🎨 Applying RTL styling...")
    style_excel(OUTPUT_FILE)
    
    print(f"\n✅ Done! File saved: {OUTPUT_FILE}")
    print("   📝 Columns for Accountant Review:")
    print("      • 'تایید حسابدار' - Approval checkbox")
    print("      • 'توضیحات اصلاحی' - Correction notes")
    print("=" * 60)


if __name__ == "__main__":
    main()
