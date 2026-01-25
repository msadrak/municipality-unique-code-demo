"""
Config Generator V7 - Reviewable Excel with Smart Clustering
=============================================================
Generates a reviewable Excel file where users can manually fix activity titles.

Key Improvements:
1. Longest Match Priority: Prefers 4-word prefixes over 2-word ones
2. Connector Awareness: Extends prefixes ending in (و, در, به, از, با, برای)
3. Raw Description Fallback: Shows raw text for manual review instead of "سایر"

Usage:
    python scripts/generate_v7_review_excel.py
"""

import pandas as pd
import re
from collections import defaultdict, Counter
from typing import Optional, List, Dict, Tuple
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_CAPITAL = "تملک دارایی سرمایه ای.xlsx"
INPUT_EXPENSE = "اعتبارات هزینه ای.xlsx"
OUTPUT_FILE = "فایل_بررسی_و_اصلاح_دستی.xlsx"

# Minimum occurrences for clustering
MIN_CLUSTER_SIZE = 2

# Persian connector words - prefixes should NOT end with these
CONNECTOR_WORDS = {"و", "در", "به", "از", "با", "برای", "های", "جهت", "روی", "تا"}

# Words to remove from descriptions (noise)
NOISE_WORDS = ["پروژه", "عملیات", "اجرای", "اجرا", "انجام", "طرح"]


# ============================================================
# CLEANING DICTIONARY (Layer 1)
# ============================================================

CLEANING_MAP = {
    # Urban & Services
    "نگهداری و توسعه فضای سبز": ["فضای سبز", "پارک", "درخت", "گیاه", "آبیاری", "چمن", "باغبانی"],
    "نظافت شهری و مدیریت پسماند": ["نظافت", "رفت و روب", "زباله", "پسماند", "جارو", "بازیافت"],
    "لایروبی انهار و مسیل‌ها": ["لایروبی", "مادی", "نهر", "کانال", "مسیل"],
    "آبرسانی و تاسیسات آبی": ["آبرسانی", "تانکر", "چاه", "قنات", "منبع آب", "شیر آتش نشانی"],
    "تاسیسات و مبلمان شهری": ["مبلمان", "نیمکت", "سطل", "تاسیسات شهری", "آذین بندی"],
    
    # Civil & Traffic
    "روکش و ترمیم آسفالت": ["آسفالت", "روکش", "لکه گیری", "قیر"],
    "پیاده‌روسازی و اصلاح معابر": ["پیاده رو", "سنگ فرش", "بلوک", "کف فرش", "زیرسازی معابر"],
    "جدول‌گذاری و کانیو": ["جدول", "کانیو", "آبراهه", "جدولگذاری"],
    "تجهیزات و علائم ترافیکی": ["ترافیک", "خط کشی", "تابلو", "سرعت گیر", "گاردریل", "چراغ راهنما"],
    "پل و تقاطع غیرهمسطح": ["پل", "زیرگذر", "روگذر", "تقاطع"],
    
    # Buildings & Facilities
    "روشنایی و نورپردازی": ["روشنایی", "نورپردازی", "برق", "لوستر", "چراغ"],
    "تعمیر و نگهداری ساختمان‌ها": ["ساختمان", "ابنیه", "تاسیسات ساختمان", "موتورخانه", "تعمیرات"],
    "احداث و بازسازی ابنیه": ["احداث", "بازسازی", "ساخت", "تکمیل"],
    "مرمت و نوسازی": ["مرمت", "نوسازی", "بهسازی", "بازآفرینی"],
    
    # Admin & HR
    "پرداخت حقوق و مزایا": ["حقوق", "دستمزد", "مزایا", "کارانه", "فوق العاده"],
    "خدمات رفاهی و انگیزشی": ["رفاهی", "پاداش", "بن", "ورزشی", "هدیه", "بیمه تکمیلی"],
    "خرید ملزومات و تجهیزات": ["خرید", "تجهیزات", "ملزومات", "اثاثیه", "لوازم"],
    "خدمات چاپ و انتشارات": ["چاپ", "نشریات", "بنر", "انتشارات"],
    
    # Financial & Legal
    "پرداخت دیون و تعهدات": ["دیون", "انتقال وجوه", "بازپرداخت", "بدهی"],
    "تملک و آزادسازی اراضی": ["تملک", "آزادسازی", "مسیر گشایی", "عرصه", "زمین"],
    "پروژه‌های مشارکتی": ["مشارکت", "سرمایه گذاری", "سرمایه‌گذاری"],
    
    # Budget & Revenue
    "مدیریت بودجه و اعتبارات": ["بودجه", "تخصیص", "موافقتنامه", "تفریغ", "اعتبار"],
    "وصول درآمد و عوارض": ["عوارض", "نوسازی", "کسب و پیشه", "درآمد", "وصول"],
    "اصفهان کارت": ["اصفهان کارت", "اصفهان‌کارت", "کارت شهروندی"],
}


# ============================================================
# TRUSTEE -> SUBSYSTEM MAPPING
# ============================================================

SUBSYSTEM_NAMES = {
    "URBAN_PLANNING": "سامانه شهرسازی",
    "CONTRACTS": "سامانه امور قراردادها",
    "PAYROLL": "سامانه حقوق و دستمزد",
    "TADAROKAT": "سامانه تدارکات",
    "BUDGET": "سامانه بودجه",
    "TREASURY": "سامانه خزانه‌داری",
    "CONTRACTORS": "سامانه امور پیمانکاران",
    "WELFARE": "سامانه رفاه کارکنان",
    "REAL_ESTATE": "سامانه املاک",
    "WAREHOUSE": "سامانه انبار و اموال",
    "REVENUE": "سامانه درآمد",
    "ISFAHAN_CARD": "سامانه اصفهان کارت",
    "INVESTMENT": "سامانه مشارکت‌ها",
    "OTHER": "سایر / عمومی",
}

TRUSTEE_TO_SUBSYSTEM = {
    "معاونت خدمات": "CONTRACTORS",
    "خدمات شهری": "CONTRACTORS",
    "معاونت فنی": "CONTRACTORS",
    "فنی عمرانی": "CONTRACTORS",
    "شهرسازی": "URBAN_PLANNING",
    "معماری": "URBAN_PLANNING",
    "معاونت مالی": "TREASURY",
    "مالی": "TREASURY",
    "خزانه": "TREASURY",
    "برنامه ریزی": "BUDGET",
    "برنامه‌ریزی": "BUDGET",
    "امور اداری": "WAREHOUSE",
    "درآمد": "REVENUE",
    "مشارکت": "INVESTMENT",
    "سرمایه گذاری": "INVESTMENT",
}


# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def clean_text(text) -> str:
    """Clean and normalize Persian text."""
    if pd.isna(text) or text is None:
        return ""
    text = str(text).strip()
    text = text.replace("ي", "ی").replace("ك", "ک")
    return text


def find_column(df: pd.DataFrame, keywords: List[str]) -> Optional[str]:
    """Find a column containing any of the keywords."""
    for col in df.columns:
        col_str = str(col).strip()
        for kw in keywords:
            if kw in col_str:
                return col
    return None


def contains_any(text: str, keywords: List[str]) -> bool:
    """Check if text contains any of the keywords."""
    if not text:
        return False
    return any(kw in text for kw in keywords)


def remove_noise(text: str) -> str:
    """Remove noise words from text."""
    for word in NOISE_WORDS:
        text = text.replace(word, "")
    # Clean up extra spaces
    text = re.sub(r'\s+', ' ', text).strip()
    return text


def extract_ngrams(text: str, n: int) -> Optional[str]:
    """Extract first N words from text."""
    # Remove numbers and special chars
    text = re.sub(r'\d+', '', text)
    text = re.sub(r'\(.*?\)', '', text)
    text = re.sub(r'[،,\-_:؛]', ' ', text)
    
    words = text.split()
    if len(words) >= n:
        return ' '.join(words[:n])
    return None


def is_valid_prefix(prefix: str) -> bool:
    """Check if prefix is valid (doesn't end with connector word)."""
    if not prefix:
        return False
    words = prefix.split()
    if not words:
        return False
    last_word = words[-1]
    # Invalid if ends with connector
    if last_word in CONNECTOR_WORDS:
        return False
    # Invalid if too short
    if len(prefix) < 5:
        return False
    return True


def extend_prefix_if_needed(prefix: str, text: str) -> str:
    """
    If prefix ends with a connector word, extend it with the next word from text.
    """
    words_prefix = prefix.split()
    if not words_prefix:
        return prefix
    
    last_word = words_prefix[-1]
    if last_word not in CONNECTOR_WORDS:
        return prefix
    
    # Find the next word in the original text
    words_text = text.split()
    prefix_len = len(words_prefix)
    
    if prefix_len < len(words_text):
        # Add next word
        extended = prefix + ' ' + words_text[prefix_len]
        return extended
    
    return prefix


def classify_to_subsystem(trustee: str, subject: str, description: str, is_capital: bool) -> str:
    """Classify row to subsystem."""
    for pattern, subsystem in TRUSTEE_TO_SUBSYSTEM.items():
        if pattern in trustee:
            return subsystem
    
    # Keywords
    if "حقوق" in subject or "دستمزد" in subject:
        return "PAYROLL"
    if "رفاهی" in description or "پاداش" in description:
        return "WELFARE"
    if "تملک" in description or "آزادسازی" in description:
        return "REAL_ESTATE"
    if "مشارکت" in description:
        return "INVESTMENT"
    if "عوارض" in description or "درآمد" in description:
        return "REVENUE"
    if "بودجه" in description or "اعتبار" in description:
        return "BUDGET"
    
    return "CONTRACTORS" if is_capital else "OTHER"


# ============================================================
# SMART CLUSTERING (V7)
# ============================================================

def smart_cluster_descriptions(descriptions: List[str]) -> Dict[str, str]:
    """
    V7 Improved Clustering with Longest Match Priority.
    Returns: {raw_description: suggested_title}
    """
    # Step 1: Generate all n-grams (2, 3, 4 words) for all descriptions
    ngram_counts = defaultdict(int)
    desc_to_ngrams = defaultdict(list)
    
    for desc in descriptions:
        cleaned = remove_noise(desc)
        for n in [4, 3, 2]:  # Try longer first
            ngram = extract_ngrams(cleaned, n)
            if ngram and len(ngram) >= 5:
                # Extend if ends with connector
                ngram = extend_prefix_if_needed(ngram, cleaned)
                if is_valid_prefix(ngram):
                    ngram_counts[ngram] += 1
                    desc_to_ngrams[desc].append((n, ngram))
    
    # Step 2: Find valid clusters (n-grams appearing >= MIN_CLUSTER_SIZE times)
    valid_clusters = {ng for ng, count in ngram_counts.items() if count >= MIN_CLUSTER_SIZE}
    
    # Step 3: Assign each description to the LONGEST matching n-gram
    result = {}
    for desc, ngrams in desc_to_ngrams.items():
        # Sort by length (longer first)
        ngrams.sort(key=lambda x: x[0], reverse=True)
        
        matched = False
        for _, ngram in ngrams:
            if ngram in valid_clusters:
                result[desc] = ngram
                matched = True
                break
        
        if not matched:
            # No cluster found - use raw description for manual review
            result[desc] = desc
    
    # Handle descriptions with no ngrams
    for desc in descriptions:
        if desc not in result:
            result[desc] = desc
    
    return result


# ============================================================
# DATA LOADING
# ============================================================

def load_excel(filepath: str, budget_type: str, filter_continuous: bool = False) -> List[dict]:
    """Load Excel file and extract rows."""
    try:
        df = pd.read_excel(filepath, engine='openpyxl')
        print(f"   ✅ Loaded: {filepath} ({len(df):,} rows)")
    except Exception as e:
        print(f"   ❌ Error: {e}")
        return []
    
    # Find columns
    desc_col = find_column(df, ['شرح ردیف', 'شرح'])
    trustee_col = find_column(df, ['متولی', 'متولي'])
    subject_col = find_column(df, ['موضوع', 'زیر موضوع'])
    row_type_col = find_column(df, ['نوع ردیف'])
    
    if not desc_col:
        print(f"   ⚠️  No description column found!")
        return []
    
    # Filter for continuous if needed
    if filter_continuous and row_type_col:
        df = df[df[row_type_col].astype(str).str.contains('مستمر', na=False)]
        print(f"   🔄 Filtered to 'مستمر': {len(df):,} rows")
    
    rows = []
    for _, row in df.iterrows():
        desc = clean_text(row.get(desc_col, ''))
        if not desc:
            continue
        
        rows.append({
            'description': desc,
            'trustee': clean_text(row.get(trustee_col, '')) if trustee_col else '',
            'subject': clean_text(row.get(subject_col, '')) if subject_col else '',
            'budget_type': budget_type
        })
    
    return rows


# ============================================================
# MAIN PROCESSING
# ============================================================

def process_all_rows(all_rows: List[dict]) -> pd.DataFrame:
    """Process all rows through the V7 pipeline."""
    
    # Group by unique description
    desc_counter = Counter([r['description'] for r in all_rows])
    
    # Build lookup for budget type and subsystem
    desc_to_info = {}
    for row in all_rows:
        desc = row['description']
        if desc not in desc_to_info:
            is_capital = row['budget_type'] == 'capital'
            subsystem = classify_to_subsystem(
                row['trustee'], row['subject'], desc, is_capital
            )
            desc_to_info[desc] = {
                'subsystem': subsystem,
                'budget_type': row['budget_type']
            }
    
    # Layer 1: Dictionary matching
    dict_matched = {}
    unmatched_descs = []
    
    for desc in desc_counter.keys():
        matched = False
        for clean_title, keywords in CLEANING_MAP.items():
            if contains_any(desc, keywords):
                dict_matched[desc] = clean_title
                matched = True
                break
        if not matched:
            unmatched_descs.append(desc)
    
    print(f"\n📊 Layer 1 (Dictionary): {len(dict_matched)} matches")
    print(f"   Remaining for clustering: {len(unmatched_descs)}")
    
    # Layer 2: Smart Clustering (V7)
    clustered = smart_cluster_descriptions(unmatched_descs)
    
    cluster_count = sum(1 for d, c in clustered.items() if d != c)
    raw_count = sum(1 for d, c in clustered.items() if d == c)
    print(f"📊 Layer 2 (Clustering): {cluster_count} clustered")
    print(f"📊 Layer 3 (Raw/Manual): {raw_count} for review")
    
    # Build output rows
    output_rows = []
    
    for desc, count in desc_counter.items():
        info = desc_to_info.get(desc, {'subsystem': 'OTHER', 'budget_type': 'unknown'})
        
        # Determine suggested title
        if desc in dict_matched:
            suggested = dict_matched[desc]
        elif desc in clustered:
            suggested = clustered[desc]
        else:
            suggested = desc
        
        # Translate budget type
        budget_type_fa = "عمرانی (سرمایه‌ای)" if info['budget_type'] == 'capital' else "جاری (هزینه‌ای)"
        
        output_rows.append({
            'نام سامانه': SUBSYSTEM_NAMES.get(info['subsystem'], 'سایر'),
            'شرح ردیف اصلی': desc,
            'عنوان پیشنهادی': suggested,
            'نوع بودجه': budget_type_fa,
            'تکرار': count
        })
    
    # Sort by subsystem and then by count
    df = pd.DataFrame(output_rows)
    df = df.sort_values(['نام سامانه', 'تکرار'], ascending=[True, False])
    
    return df


def style_excel(filepath: str):
    """Apply styling to Excel file."""
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    editable_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style header
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Style data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for col_idx, cell in enumerate(row, start=1):
            cell.alignment = cell_align
            cell.border = thin_border
            # Highlight editable column (عنوان پیشنهادی = column 3)
            if col_idx == 3:
                cell.fill = editable_fill
    
    # Column widths
    column_widths = {
        'A': 28,  # نام سامانه
        'B': 60,  # شرح ردیف اصلی
        'C': 45,  # عنوان پیشنهادی (EDITABLE)
        'D': 20,  # نوع بودجه
        'E': 10,  # تکرار
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Row heights
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 25
    
    # Freeze header and RTL
    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True
    
    wb.save(filepath)


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("📊 CONFIG GENERATOR V7 - REVIEWABLE EXCEL")
    print("=" * 70)
    print()
    print("V7 Improvements:")
    print("  • Longest Match Priority (4-word prefixes before 2-word)")
    print("  • Connector Awareness (extends 'مرمت و' to 'مرمت و نوسازی')")
    print("  • Raw Fallback for Manual Review (no 'سایر' hiding)")
    print()
    
    # Load data
    print("📁 Loading Excel files...")
    capital_rows = load_excel(INPUT_CAPITAL, 'capital', filter_continuous=True)
    expense_rows = load_excel(INPUT_EXPENSE, 'expense', filter_continuous=False)
    
    all_rows = capital_rows + expense_rows
    print(f"\n   Total rows: {len(all_rows):,}")
    
    # Process
    print("\n🔄 Processing through V7 Pipeline...")
    df = process_all_rows(all_rows)
    
    # Save
    print(f"\n💾 Saving to: {OUTPUT_FILE}")
    df.to_excel(OUTPUT_FILE, index=False, engine='openpyxl')
    
    # Style
    print("🎨 Applying styling...")
    style_excel(OUTPUT_FILE)
    
    # Summary
    print("\n" + "=" * 70)
    print("📋 SUMMARY")
    print("=" * 70)
    print(f"   Output: {OUTPUT_FILE}")
    print(f"   Unique Descriptions: {len(df)}")
    print()
    print("   📝 INSTRUCTIONS FOR REVIEWER:")
    print("   1. Open the Excel file")
    print("   2. Review the 'عنوان پیشنهادی' column (highlighted in YELLOW)")
    print("   3. Edit titles that are incorrect or too long")
    print("   4. Save the file for import into the final config")
    print("=" * 70)


if __name__ == "__main__":
    main()
