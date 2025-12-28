"""
Config Generator V8 - Strict Review with Color Filter
=======================================================
Generates a reviewable Excel file with "Double-Lock" filtering:
1. Text Condition: نوع ردیف == "مستمر"
2. Color Condition: Row background is WHITE or NO FILL (excludes highlighted rows)

Uses openpyxl directly to read cell colors.

Usage:
    python scripts/generate_v8_strict_review.py
"""

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import re
from collections import defaultdict, Counter
from typing import Optional, List, Dict, Tuple, Set


# ============================================================
# CONFIGURATION
# ============================================================

INPUT_CAPITAL = "تملک دارایی سرمایه ای.xlsx"
INPUT_EXPENSE = "اعتبارات هزینه ای.xlsx"
OUTPUT_FILE = "فایل_بررسی_نهایی_V8_با_فیلتر_رنگ.xlsx"

# Clustering settings
MIN_CLUSTER_SIZE = 3
MIN_PREFIX_WORDS = 3

# Persian connector words
CONNECTOR_WORDS = {"و", "در", "به", "از", "با", "برای", "های", "جهت", "روی", "تا"}

# Valid colors (White or No Fill)
# No Fill: index = '00000000', rgb = None, theme = None
# White: index = 'FFFFFFFF' or RGB = 'FFFFFF'
VALID_COLOR_INDICES = {'00000000', '00000000', 0, None}
VALID_COLOR_RGB = {'FFFFFFFF', 'FFFFFF', 'ffffff', None, '00000000'}


# ============================================================
# CLEANING DICTIONARY (Layer 1)
# ============================================================

CLEANING_MAP = {
    # Urban & Services
    "نگهداری و توسعه فضای سبز": ["فضای سبز", "پارک", "درخت", "گیاه", "آبیاری", "چمن", "باغبانی"],
    "نظافت شهری و مدیریت پسماند": ["نظافت", "رفت و روب", "زباله", "پسماند", "جارو", "بازیافت"],
    "لایروبی انهار و مسیل‌ها": ["لایروبی", "مادی", "نهر", "کانال", "مسیل"],
    "آبرسانی و تاسیسات آبی": ["آبرسانی", "تانکر", "چاه", "قنات", "منبع آب", "شیر آتش نشانی"],
    "تاسیسات و مبلمان شهری": ["مبلمان", "نیمکت", "سطل", "تاسیسات شهری", "آذین بندی"],
    
    # Civil & Traffic
    "روکش و ترمیم آسفالت": ["آسفالت", "روکش", "لکه گیری", "قیر"],
    "پیاده‌روسازی و اصلاح معابر": ["پیاده رو", "سنگ فرش", "بلوک", "کف فرش", "زیرسازی معابر", "زیرسازی"],
    "جدول‌گذاری و کانیو": ["جدول", "کانیو", "آبراهه", "جدولگذاری"],
    "تجهیزات و علائم ترافیکی": ["ترافیک", "خط کشی", "تابلو", "سرعت گیر", "گاردریل", "چراغ راهنما"],
    "پل و تقاطع غیرهمسطح": ["پل", "زیرگذر", "روگذر", "تقاطع"],
    
    # Buildings & Facilities
    "روشنایی و نورپردازی": ["روشنایی", "نورپردازی", "برق", "لوستر", "چراغ"],
    "تعمیر و نگهداری ساختمان‌ها": ["ساختمان", "ابنیه", "تاسیسات ساختمان", "موتورخانه", "تعمیرات"],
    "احداث و بازسازی ابنیه": ["احداث", "بازسازی", "ساخت", "تکمیل"],
    "مرمت و نوسازی": ["مرمت", "نوسازی", "بهسازی", "بازآفرینی"],
    
    # Admin & HR
    "پرداخت حقوق و مزایا": ["حقوق", "دستمزد", "مزایا", "کارانه", "فوق العاده"],
    "خدمات رفاهی و انگیزشی": ["رفاهی", "پاداش", "بن", "ورزشی", "هدیه", "بیمه تکمیلی"],
    "خرید ملزومات و تجهیزات": ["خرید", "تجهیزات", "ملزومات", "اثاثیه", "لوازم"],
    "خدمات چاپ و انتشارات": ["چاپ", "نشریات", "بنر", "انتشارات"],
    
    # Financial & Legal
    "پرداخت دیون و تعهدات": ["دیون", "انتقال وجوه", "بازپرداخت", "بدهی"],
    "تملک و آزادسازی اراضی": ["تملک", "آزادسازی", "مسیر گشایی", "عرصه", "زمین"],
    "پروژه‌های مشارکتی": ["مشارکت", "سرمایه گذاری", "سرمایه‌گذاری"],
    
    # Budget & Revenue
    "مدیریت بودجه و اعتبارات": ["بودجه", "تخصیص", "موافقتنامه", "تفریغ", "اعتبار"],
    "وصول درآمد و عوارض": ["عوارض", "نوسازی", "کسب و پیشه", "درآمد", "وصول"],
    "اصفهان کارت": ["اصفهان کارت", "اصفهان‌کارت", "کارت شهروندی"],
}


# ============================================================
# SUBSYSTEM MAPPING
# ============================================================

SUBSYSTEM_NAMES = {
    "URBAN_PLANNING": "سامانه شهرسازی",
    "CONTRACTS": "سامانه امور قراردادها",
    "PAYROLL": "سامانه حقوق و دستمزد",
    "TADAROKAT": "سامانه تدارکات",
    "BUDGET": "سامانه بودجه",
    "TREASURY": "سامانه خزانه‌داری",
    "CONTRACTORS": "سامانه امور پیمانکاران",
    "WELFARE": "سامانه رفاه کارکنان",
    "REAL_ESTATE": "سامانه املاک",
    "WAREHOUSE": "سامانه انبار و اموال",
    "REVENUE": "سامانه درآمد",
    "ISFAHAN_CARD": "سامانه اصفهان کارت",
    "INVESTMENT": "سامانه مشارکت‌ها",
    "OTHER": "سایر / عمومی",
}

TRUSTEE_TO_SUBSYSTEM = {
    "معاونت خدمات": "CONTRACTORS",
    "خدمات شهری": "CONTRACTORS",
    "معاونت فنی": "CONTRACTORS",
    "فنی عمرانی": "CONTRACTORS",
    "شهرسازی": "URBAN_PLANNING",
    "معماری": "URBAN_PLANNING",
    "معاونت مالی": "TREASURY",
    "مالی": "TREASURY",
    "خزانه": "TREASURY",
    "برنامه ریزی": "BUDGET",
    "برنامه‌ریزی": "BUDGET",
    "امور اداری": "WAREHOUSE",
    "درآمد": "REVENUE",
    "مشارکت": "INVESTMENT",
    "سرمایه گذاری": "INVESTMENT",
}


# ============================================================
# COLOR CHECKING UTILITIES
# ============================================================

def is_white_or_no_fill(cell) -> bool:
    """
    Check if a cell has WHITE or NO FILL background.
    Returns True if valid (white/transparent), False if colored.
    """
    try:
        fill = cell.fill
        
        # No fill case
        if fill is None:
            return True
        
        # Check fill type
        fill_type = fill.fill_type
        if fill_type is None or fill_type == 'none':
            return True
        
        # Check solid fill colors
        if fill_type == 'solid':
            fg_color = fill.fgColor
            
            if fg_color is None:
                return True
            
            # Check if it's a theme color (usually means no explicit color)
            if fg_color.type == 'theme':
                # Theme 0 is usually white/background
                if fg_color.theme == 0:
                    return True
                # Other themes might be colored
                return False
            
            # Check RGB value
            if fg_color.type == 'rgb':
                rgb = fg_color.rgb
                if rgb is None:
                    return True
                rgb_str = str(rgb).upper()
                # White: FFFFFFFF or 00000000 (transparent)
                if rgb_str in ('FFFFFFFF', '00000000', 'FFFFFF'):
                    return True
                # Check if it's a light color (near white)
                if rgb_str.startswith('FF') and len(rgb_str) == 8:
                    # Extract RGB values
                    r = int(rgb_str[2:4], 16)
                    g = int(rgb_str[4:6], 16)
                    b = int(rgb_str[6:8], 16)
                    # If very close to white (>250 each), accept as white
                    if r > 250 and g > 250 and b > 250:
                        return True
                return False
            
            # Check indexed color
            if fg_color.type == 'indexed':
                idx = fg_color.indexed
                # Index 0 and 64 are typically black/automatic, 
                # but for background 0 often means no fill
                if idx in (0, 64, None):
                    return True
                return False
        
        return True  # Default to accepting
        
    except Exception as e:
        # If any error, assume it's valid
        return True


def find_column_index(header_row, keywords: List[str]) -> Optional[int]:
    """Find column index (1-based) that contains any keyword."""
    for idx, cell in enumerate(header_row, start=1):
        cell_value = str(cell.value or "").strip()
        for kw in keywords:
            if kw in cell_value:
                return idx
    return None


# ============================================================
# TEXT UTILITIES
# ============================================================

def clean_text(text) -> str:
    """Clean and normalize Persian text."""
    if text is None:
        return ""
    text = str(text).strip()
    text = text.replace("ي", "ی").replace("ك", "ک")
    return text


def contains_any(text: str, keywords: List[str]) -> bool:
    """Check if text contains any of the keywords."""
    if not text:
        return False
    return any(kw in text for kw in keywords)


def extract_prefix(text: str, n: int) -> Optional[str]:
    """Extract first N words as prefix."""
    # Remove noise
    text = re.sub(r'\d+', '', text)
    text = re.sub(r'\(.*?\)', '', text)
    text = re.sub(r'[،,\-_:؛]', ' ', text)
    
    words = text.split()
    if len(words) >= n:
        prefix = ' '.join(words[:n])
        # Don't accept if ends with connector
        if words[n-1] in CONNECTOR_WORDS:
            if len(words) > n:
                prefix = ' '.join(words[:n+1])
            else:
                return None
        return prefix if len(prefix) >= 6 else None
    return None


def classify_to_subsystem(trustee: str, description: str, is_capital: bool) -> str:
    """Classify row to subsystem."""
    for pattern, subsystem in TRUSTEE_TO_SUBSYSTEM.items():
        if pattern in trustee:
            return subsystem
    
    if "رفاهی" in description or "پاداش" in description:
        return "WELFARE"
    if "حقوق" in description or "دستمزد" in description:
        return "PAYROLL"
    if "تملک" in description:
        return "REAL_ESTATE"
    if "مشارکت" in description:
        return "INVESTMENT"
    if "عوارض" in description or "درآمد" in description:
        return "REVENUE"
    if "بودجه" in description:
        return "BUDGET"
    
    return "CONTRACTORS" if is_capital else "OTHER"


# ============================================================
# DATA LOADING WITH COLOR FILTER
# ============================================================

def load_capital_with_color_filter(filepath: str) -> List[dict]:
    """
    Load capital budget file with DOUBLE-LOCK filter:
    1. نوع ردیف == "مستمر"
    2. Row background is WHITE or NO FILL
    """
    print(f"   📂 Loading: {filepath}")
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"   ❌ Error loading file: {e}")
        return []
    
    # Get header row (row 1)
    header_row = list(ws[1])
    
    # Find columns
    desc_col = find_column_index(header_row, ['شرح ردیف', 'شرح'])
    type_col = find_column_index(header_row, ['نوع ردیف'])
    trustee_col = find_column_index(header_row, ['متولی', 'متولي'])
    
    if not desc_col:
        print(f"   ⚠️  No description column found!")
        return []
    
    print(f"   📊 Columns: desc={desc_col}, type={type_col}, trustee={trustee_col}")
    
    total_rows = 0
    text_filtered = 0
    color_filtered = 0
    valid_rows = []
    
    # Iterate through data rows (starting from row 2)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        total_rows += 1
        
        # Get cells
        desc_cell = row[desc_col - 1] if desc_col else None
        type_cell = row[type_col - 1] if type_col else None
        trustee_cell = row[trustee_col - 1] if trustee_col else None
        
        # Get values
        desc_value = clean_text(desc_cell.value) if desc_cell else ""
        type_value = clean_text(type_cell.value) if type_cell else ""
        trustee_value = clean_text(trustee_cell.value) if trustee_cell else ""
        
        if not desc_value:
            continue
        
        # FILTER 1: Text condition (نوع ردیف == مستمر)
        if "مستمر" not in type_value:
            text_filtered += 1
            continue
        
        # FILTER 2: Color condition (White or No Fill)
        # Check the first cell in the row as indicator
        first_cell = row[0]
        if not is_white_or_no_fill(first_cell):
            color_filtered += 1
            continue
        
        # Row passed both filters
        valid_rows.append({
            'description': desc_value,
            'trustee': trustee_value,
            'budget_type': 'capital'
        })
    
    wb.close()
    
    print(f"   📈 Total rows: {total_rows}")
    print(f"   🚫 Filtered by text (not مستمر): {text_filtered}")
    print(f"   🎨 Filtered by color (highlighted): {color_filtered}")
    print(f"   ✅ Valid rows: {len(valid_rows)}")
    
    return valid_rows


def load_expense_file(filepath: str) -> List[dict]:
    """Load expense budget file (all rows are valid)."""
    print(f"   📂 Loading: {filepath}")
    
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as e:
        print(f"   ❌ Error: {e}")
        return []
    
    # Get header row
    header_row = list(ws[1])
    
    # Find columns
    desc_col = find_column_index(header_row, ['شرح ردیف', 'شرح'])
    trustee_col = find_column_index(header_row, ['متولی', 'متولي'])
    
    if not desc_col:
        print(f"   ⚠️  No description column found!")
        return []
    
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        desc_cell = row[desc_col - 1] if desc_col else None
        trustee_cell = row[trustee_col - 1] if trustee_col else None
        
        desc_value = clean_text(desc_cell.value) if desc_cell else ""
        trustee_value = clean_text(trustee_cell.value) if trustee_cell else ""
        
        if desc_value:
            rows.append({
                'description': desc_value,
                'trustee': trustee_value,
                'budget_type': 'expense'
            })
    
    wb.close()
    print(f"   ✅ Loaded: {len(rows)} rows")
    
    return rows


# ============================================================
# CLEANING PIPELINE
# ============================================================

def apply_dictionary(description: str) -> Optional[str]:
    """Layer 1: Check against CLEANING_MAP."""
    for clean_title, keywords in CLEANING_MAP.items():
        if contains_any(description, keywords):
            return clean_title
    return None


def build_clusters(descriptions: List[str]) -> Dict[str, str]:
    """Layer 2: Build clusters from common prefixes (longest first)."""
    # Count prefixes of different lengths
    prefix_counts = Counter()
    desc_to_prefixes = defaultdict(list)
    
    for desc in descriptions:
        for n in [4, 3]:  # Try 4-word and 3-word prefixes
            prefix = extract_prefix(desc, n)
            if prefix:
                prefix_counts[prefix] += 1
                desc_to_prefixes[desc].append((n, prefix))
    
    # Find valid clusters
    valid_clusters = {p for p, c in prefix_counts.items() if c >= MIN_CLUSTER_SIZE}
    
    # Assign each description to longest matching cluster
    result = {}
    for desc, prefixes in desc_to_prefixes.items():
        prefixes.sort(key=lambda x: x[0], reverse=True)  # Longest first
        matched = False
        for _, prefix in prefixes:
            if prefix in valid_clusters:
                result[desc] = prefix
                matched = True
                break
        if not matched:
            result[desc] = desc  # Raw fallback
    
    # Handle descriptions with no prefixes
    for desc in descriptions:
        if desc not in result:
            result[desc] = desc
    
    return result


def process_all_rows(all_rows: List[dict]) -> List[dict]:
    """Process all rows through the cleaning pipeline."""
    
    # Count unique descriptions
    desc_counter = Counter(r['description'] for r in all_rows)
    
    # Build lookup for metadata
    desc_to_meta = {}
    for row in all_rows:
        desc = row['description']
        if desc not in desc_to_meta:
            is_capital = row['budget_type'] == 'capital'
            subsystem = classify_to_subsystem(row['trustee'], desc, is_capital)
            desc_to_meta[desc] = {
                'subsystem': subsystem,
                'budget_type': row['budget_type']
            }
    
    # Layer 1: Dictionary matching
    dict_matched = {}
    unmatched = []
    
    for desc in desc_counter.keys():
        title = apply_dictionary(desc)
        if title:
            dict_matched[desc] = title
        else:
            unmatched.append(desc)
    
    print(f"\n📊 Layer 1 (Dictionary): {len(dict_matched)} matched")
    print(f"   Remaining: {len(unmatched)}")
    
    # Layer 2: Clustering
    clustered = build_clusters(unmatched)
    
    cluster_count = sum(1 for d, c in clustered.items() if d != c)
    raw_count = sum(1 for d, c in clustered.items() if d == c)
    print(f"📊 Layer 2 (Clustering): {cluster_count} clustered")
    print(f"📊 Layer 3 (Raw/Manual): {raw_count} for review")
    
    # Build output
    output = []
    for desc, count in desc_counter.items():
        meta = desc_to_meta.get(desc, {'subsystem': 'OTHER', 'budget_type': 'unknown'})
        
        if desc in dict_matched:
            suggested = dict_matched[desc]
        elif desc in clustered:
            suggested = clustered[desc]
        else:
            suggested = desc
        
        budget_fa = "عمرانی (سرمایه‌ای)" if meta['budget_type'] == 'capital' else "جاری (هزینه‌ای)"
        
        output.append({
            'سامانه': SUBSYSTEM_NAMES.get(meta['subsystem'], 'سایر'),
            'شرح_اصلی': desc,
            'عنوان_پیشنهادی': suggested,
            'نوع_بودجه': budget_fa,
            'تکرار': count
        })
    
    # Sort by subsystem
    output.sort(key=lambda x: x['سامانه'])
    
    return output


# ============================================================
# EXCEL OUTPUT
# ============================================================

def write_output_excel(data: List[dict], filepath: str):
    """Write data to styled Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "بررسی فعالیت‌ها"
    
    # Headers
    headers = ['سامانه', 'شرح_اصلی', 'عنوان_پیشنهادی', 'نوع_بودجه', 'تکرار']
    
    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    editable_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Write data
    for row_idx, item in enumerate(data, start=2):
        for col_idx, key in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=item.get(key, ''))
            cell.alignment = cell_align
            cell.border = thin_border
            # Highlight editable column (عنوان_پیشنهادی = column 3)
            if col_idx == 3:
                cell.fill = editable_fill
    
    # Column widths
    widths = {'A': 28, 'B': 55, 'C': 40, 'D': 20, 'E': 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Row heights
    ws.row_dimensions[1].height = 30
    for row_idx in range(2, len(data) + 2):
        ws.row_dimensions[row_idx].height = 24
    
    # Freeze header and RTL
    ws.freeze_panes = "A2"
    ws.sheet_view.rightToLeft = True
    
    wb.save(filepath)


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 70)
    print("📊 CONFIG GENERATOR V8 - STRICT REVIEW WITH COLOR FILTER")
    print("=" * 70)
    print()
    print("Double-Lock Filter:")
    print("  1. Text Condition: نوع ردیف == 'مستمر'")
    print("  2. Color Condition: Row background is WHITE or NO FILL")
    print()
    
    # Load Capital with color filter
    print("📁 Loading Capital Budget (with color filter)...")
    capital_rows = load_capital_with_color_filter(INPUT_CAPITAL)
    
    # Load Expense (all valid)
    print("\n📁 Loading Expense Budget...")
    expense_rows = load_expense_file(INPUT_EXPENSE)
    
    # Combine
    all_rows = capital_rows + expense_rows
    print(f"\n📊 Total valid rows: {len(all_rows)}")
    
    # Process
    print("\n🔄 Processing through Cleaning Pipeline...")
    output_data = process_all_rows(all_rows)
    
    # Write output
    print(f"\n💾 Writing to: {OUTPUT_FILE}")
    write_output_excel(output_data, OUTPUT_FILE)
    
    # Summary
    print("\n" + "=" * 70)
    print("📋 SUMMARY")
    print("=" * 70)
    print(f"   Output File: {OUTPUT_FILE}")
    print(f"   Unique Descriptions: {len(output_data)}")
    print()
    print("   📝 REVIEW INSTRUCTIONS:")
    print("   1. Open the Excel file")
    print("   2. Review 'عنوان_پیشنهادی' column (YELLOW)")
    print("   3. Edit incorrect or overly long titles")
    print("   4. Save for final config import")
    print("=" * 70)


if __name__ == "__main__":
    main()
