"""
Mass Contractor Import Script
===============================

Reads contractor names from the xlsx file and bulk-inserts them
into the Contractor table with placeholder national_ids (LEGACY-XXXX).

Usage:
    python scripts/import_mass_contractors.py
"""

import sys
import os
import time

# Allow imports from project root
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook
from app.database import SessionLocal
from app.models import Contractor


XLSX_PATH = os.path.join(
    os.path.dirname(__file__), "..", "لیست پیماکاران.xlsx"
)

# Placeholder national_id prefix
LEGACY_PREFIX = "LEGACY"

# How many rows to commit per batch
BATCH_SIZE = 500


def read_names_from_excel(path: str) -> list[str]:
    """Read contractor names from column A of the first sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active sheet found in the workbook.")

    names: list[str] = []
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        cell_value = row[0]
        if cell_value is None:
            continue
        name = str(cell_value).strip()
        if name:
            names.append(name)

    wb.close()

    # Auto-detect header: skip first row if it looks like a header
    if names and names[0] in ("نام", "نام پیمانکار", "name", "Name", "شرکت"):
        print(f"  ℹ  Detected header row: '{names[0]}' — skipping it.")
        names = names[1:]

    return names


def import_contractors():
    """Main import logic."""
    print("=" * 60)
    print("  Mass Contractor Import")
    print("=" * 60)

    # ── 1. Read Excel ──────────────────────────────────────────
    abs_path = os.path.abspath(XLSX_PATH)
    print(f"\n📂 Reading: {abs_path}")

    if not os.path.isfile(abs_path):
        print(f"  ✗  File not found: {abs_path}")
        sys.exit(1)

    names = read_names_from_excel(abs_path)
    print(f"  ✓  Read {len(names)} names from Excel.")

    if not names:
        print("  ⚠  No names found. Nothing to import.")
        return

    # ── 2. De-duplicate within the Excel list itself ───────────
    seen: set[str] = set()
    unique_names: list[str] = []
    excel_dupes = 0
    for name in names:
        key = name.lower()
        if key in seen:
            excel_dupes += 1
            continue
        seen.add(key)
        unique_names.append(name)

    if excel_dupes:
        print(f"  ℹ  Removed {excel_dupes} duplicate names within Excel.")

    # ── 3. Check existing names in DB ──────────────────────────
    db = SessionLocal()
    try:
        existing_names: set[str] = {
            row[0].lower()
            for row in db.query(Contractor.company_name).all()
            if row[0]
        }
        print(f"  ℹ  Found {len(existing_names)} existing contractors in DB.")

        # Find the highest existing LEGACY-XXXX number to continue from
        existing_legacy = (
            db.query(Contractor.national_id)
            .filter(Contractor.national_id.like(f"{LEGACY_PREFIX}-%"))
            .all()
        )
        max_seq = 0
        for (nid,) in existing_legacy:
            try:
                seq = int(nid.split("-", 1)[1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass

        # ── 4. Build new Contractor objects ────────────────────
        to_insert: list[Contractor] = []
        skipped_db = 0
        seq_counter = max_seq

        for name in unique_names:
            if name.lower() in existing_names:
                skipped_db += 1
                continue

            seq_counter += 1
            national_id = f"{LEGACY_PREFIX}-{seq_counter:04d}"

            to_insert.append(
                Contractor(
                    national_id=national_id,
                    company_name=name,
                    is_verified=True,
                    source_system="LEGACY_IMPORT",
                )
            )

        print(f"  ℹ  Skipped {skipped_db} names (already in DB).")
        print(f"  ℹ  Preparing to insert {len(to_insert)} new contractors.")

        if not to_insert:
            print("\n✅ Nothing to insert — all names already exist.")
            return

        # ── 5. Bulk insert in batches ──────────────────────────
        t0 = time.time()
        for i in range(0, len(to_insert), BATCH_SIZE):
            batch = to_insert[i : i + BATCH_SIZE]
            db.bulk_save_objects(batch)
            db.commit()
            print(f"  → Committed batch {i // BATCH_SIZE + 1} "
                  f"({len(batch)} records)")

        elapsed = time.time() - t0

        # ── 6. Summary ────────────────────────────────────────
        print()
        print("=" * 60)
        print(f"  ✅ Import complete in {elapsed:.1f}s")
        print(f"     Total read from Excel : {len(names)}")
        print(f"     Duplicates in Excel   : {excel_dupes}")
        print(f"     Skipped (in DB)       : {skipped_db}")
        print(f"     Inserted              : {len(to_insert)}")
        print(f"     ID range              : {LEGACY_PREFIX}-{max_seq + 1:04d} → "
              f"{LEGACY_PREFIX}-{max_seq + len(to_insert):04d}")
        print("=" * 60)

    finally:
        db.close()


if __name__ == "__main__":
    import_contractors()
