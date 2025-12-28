"""
Export Config to Excel for Accountant Verification
====================================================
Converts app/config/config_master.json to a human-readable Excel report.

Usage:
    python scripts/export_config_to_excel.py
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# ============================================================
# CONFIGURATION
# ============================================================

INPUT_FILE = "app/config/config_master.json"
OUTPUT_FILE = "Config_V4_Report.xlsx"

# Translation mappings
BUDGET_TYPE_TRANSLATIONS = {
    "capital": "عمرانی/سرمایه‌ای",
    "expense": "هزینه‌ای/جاری",
    "both": "هر دو نوع"
}

FREQUENCY_TRANSLATIONS = {
    "DAILY": "روزانه",
    "MONTHLY": "ماهانه",
    "YEARLY": "سالانه",
    None: "-"
}

# ============================================================
# DATA PROCESSING
# ============================================================

def load_config() -> dict:
    """Load the JSON config file."""
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def extract_budget_type(constraints: list) -> str:
    """Extract allowed budget types from constraints list."""
    if not constraints:
        return "نامشخص"
    
    budget_types = set()
    for constraint in constraints:
        allowed = constraint.get("allowed_budget_types", [])
        if allowed:
            budget_types.update(allowed)
    
    if not budget_types:
        return "نامشخص"
    
    if "capital" in budget_types and "expense" in budget_types:
        return BUDGET_TYPE_TRANSLATIONS["both"]
    elif "capital" in budget_types:
        return BUDGET_TYPE_TRANSLATIONS["capital"]
    elif "expense" in budget_types:
        return BUDGET_TYPE_TRANSLATIONS["expense"]
    else:
        return "نامشخص"


def flatten_config(config: dict) -> list:
    """Flatten the nested JSON into a list of rows."""
    rows = []
    
    for subsystem in config.get("subsystems", []):
        system_name = subsystem.get("title", "")
        system_code = subsystem.get("code", "")
        
        for activity in subsystem.get("activities", []):
            activity_code = activity.get("code", "")
            activity_title = activity.get("title", "")
            frequency = activity.get("frequency")
            constraints = activity.get("constraints", [])
            
            budget_type = extract_budget_type(constraints)
            frequency_persian = FREQUENCY_TRANSLATIONS.get(frequency, frequency or "-")
            
            rows.append({
                "ردیف": len(rows) + 1,
                "نام سامانه": system_name,
                "کد سامانه": system_code,
                "کد فعالیت": activity_code,
                "عنوان فعالیت": activity_title,
                "نوع بودجه مجاز": budget_type,
                "دوره": frequency_persian,
                "تایید حسابدار": "",  # Blank for manual input
                "توضیحات": ""  # Blank for manual input
            })
    
    return rows


# ============================================================
# EXCEL FORMATTING
# ============================================================

def style_excel(output_path: str):
    """Apply professional styling to the Excel file."""
    wb = load_workbook(output_path)
    ws = wb.active
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Alternating row colors
    light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Style data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.alignment = cell_alignment
            cell.border = thin_border
            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = light_fill
    
    # Set column widths (RTL-friendly)
    column_widths = {
        'A': 6,   # ردیف
        'B': 30,  # نام سامانه
        'C': 18,  # کد سامانه
        'D': 20,  # کد فعالیت
        'E': 45,  # عنوان فعالیت
        'F': 18,  # نوع بودجه مجاز
        'G': 12,  # دوره
        'H': 15,  # تایید حسابدار
        'I': 30   # توضیحات
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height
    ws.row_dimensions[1].height = 30  # Header
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 25
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Set sheet direction to RTL
    ws.sheet_view.rightToLeft = True
    
    wb.save(output_path)


# ============================================================
# MAIN
# ============================================================

def main():
    """Main entry point."""
    print("=" * 60)
    print("📊 EXPORT CONFIG TO EXCEL")
    print("=" * 60)
    
    # Load config
    print(f"\n📁 Loading: {INPUT_FILE}")
    config = load_config()
    
    # Flatten data
    print("🔄 Flattening data...")
    rows = flatten_config(config)
    
    # Create DataFrame
    df = pd.DataFrame(rows)
    
    # Calculate summary
    print(f"\n📋 Summary:")
    print(f"   Total subsystems: {len(config.get('subsystems', []))}")
    print(f"   Total activities: {len(rows)}")
    
    # Group by subsystem
    print("\n   Activities per subsystem:")
    for subsystem in config.get("subsystems", []):
        act_count = len(subsystem.get("activities", []))
        print(f"   - {subsystem.get('title', '')}: {act_count}")
    
    # Export to Excel
    print(f"\n💾 Exporting to: {OUTPUT_FILE}")
    df.to_excel(OUTPUT_FILE, index=False, engine='openpyxl')
    
    # Apply styling
    print("🎨 Applying styling...")
    style_excel(OUTPUT_FILE)
    
    print(f"\n✅ Done! File saved: {OUTPUT_FILE}")
    print("   📝 The accountant can now review and fill in:")
    print("      - 'تایید حسابدار' column (approval)")
    print("      - 'توضیحات' column (comments)")
    print("=" * 60)


if __name__ == "__main__":
    main()
